import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import zipfile
from copy import copy

st.set_page_config(page_title="Mentor-wise Splitter (Formatted)", layout="wide")
st.title("📚 Mentor-wise Student Splitter and Link Generator")

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

def df_to_formatted_workbook(original_wb, original_ws_name, df_filtered, id_col_name="ADEK Applicant ID"):
    """
    Create a new workbook and copy styles/hyperlinks from the original workbook
    by looking up the original row for each student using the id_col_name.
    """
    original_ws = original_wb[original_ws_name]

    # Build a map: value_of_id_col -> original_row_number
    id_col_idx = None
    header_row = 1
    max_col = original_ws.max_column
    # find header column index for the id column
    for c in range(1, max_col + 1):
        val = original_ws.cell(row=header_row, column=c).value
        if val == id_col_name:
            id_col_idx = c
            break
    if id_col_idx is None:
        raise ValueError(f"ID column '{id_col_name}' not found in original sheet headers.")

    id_to_row = {}
    for r in range(2, original_ws.max_row + 1):
        key = original_ws.cell(row=r, column=id_col_idx).value
        if key is not None:
            id_to_row[str(key)] = r  # string keys to be safe

    # Create new workbook
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Students"

    # Copy column widths
    for col_letter, dim in original_ws.column_dimensions.items():
        try:
            new_wb.active.column_dimensions[col_letter].width = dim.width
        except Exception:
            pass  # ignore any columns that can't be set

    # Copy row heights for header and up to number of rows we will write
    try:
        # copy header row height
        new_ws.row_dimensions[1].height = original_ws.row_dimensions.get(1).height
    except Exception:
        pass

    # Write headers copying formatting from original header row
    cols = list(df_filtered.columns)
    for col_idx, col_name in enumerate(cols, start=1):
        src = original_ws.cell(row=1, column=col_idx)
        tgt = new_ws.cell(row=1, column=col_idx, value=src.value if src.value is not None else col_name)

        # copy style objects safely using copy()
        try:
            tgt.font = copy(src.font)
            tgt.fill = copy(src.fill)
            tgt.border = copy(src.border)
            tgt.alignment = copy(src.alignment)
            tgt.number_format = src.number_format
        except Exception:
            # fallback: ignore style copy failures
            pass

        # preserve hyperlink if any on header
        if src.hyperlink:
            # src.hyperlink may be Hyperlink object or str
            link = getattr(src.hyperlink, "target", src.hyperlink)
            tgt.hyperlink = link
            try:
                tgt.style = "Hyperlink"
            except Exception:
                pass

    # For each row in filtered df, find corresponding original row and copy cell-by-cell
    row_out = 2
    for _, row in df_filtered.iterrows():
        # find matching original row using id column
        id_val = row.get(id_col_name)
        orig_row_num = id_to_row.get(str(id_val))
        # If not found, fallback to using a template row (row 2) from original
        if orig_row_num is None:
            orig_row_num = 2 if original_ws.max_row >= 2 else 1

        # copy row height if present
        try:
            h = original_ws.row_dimensions.get(orig_row_num)
            if h and h.height:
                new_ws.row_dimensions[row_out].height = h.height
        except Exception:
            pass

        for col_idx, col_name in enumerate(cols, start=1):
            val = row[col_name]
            src = original_ws.cell(row=orig_row_num, column=col_idx)
            tgt = new_ws.cell(row=row_out, column=col_idx, value=val)

            # copy style objects safely using copy()
            try:
                tgt.font = copy(src.font)
                tgt.fill = copy(src.fill)
                tgt.border = copy(src.border)
                tgt.alignment = copy(src.alignment)
                tgt.number_format = src.number_format
            except Exception:
                pass

            # preserve hyperlink: if the original cell had a hyperlink, reuse it;
            # otherwise if the value itself is a hyperlink text, set that as hyperlink
            try:
              if col_name == "Microsoft Form Link":  
                if src.hyperlink:
                    link = getattr(src.hyperlink, "target", src.hyperlink)
                elif isinstance(val, str) and (val.startswith("http://") or val.startswith("https://")):
                    link = val
                else:
                    link = None

                if link:
                    adek_id = str(row.get(id_col_name, ""))  
                    tgt.value = adek_id  
                    tgt.hyperlink = link
                    try:
                        tgt.style = "Hyperlink"
                    except:
                        pass
                else:
                    tgt.value = val

              else:
                  pass
              
            except Exception:
                # ignore hyperlink copy errors
                pass

        row_out += 1

    return new_wb


if uploaded_file:
    # read with pandas for filtering UI convenience
    df = pd.read_excel(uploaded_file)

    if "Current Mentor" not in df.columns or "Team Lead" not in df.columns:
        st.error("❌ Required columns missing! Make sure 'Current Mentor' and 'Team Lead' exist.")
        st.stop()

    st.success("✅ File uploaded successfully!")

    # load original workbook to copy styles/hyperlinks from
    original_wb = load_workbook(uploaded_file, data_only=False)
    sheet_name = original_wb.sheetnames[0]

    # UI dropdowns
    team_leads = ["All"] + sorted(df["Team Lead"].dropna().unique())
    selected_team_lead = st.selectbox("Select Team Lead", team_leads)

    if selected_team_lead != "All":
        filtered_df = df[df["Team Lead"] == selected_team_lead]
    else:
        filtered_df = df.copy()

    mentors = sorted(filtered_df["Current Mentor"].dropna().unique())
    selected_mentor = st.selectbox("Select Mentor", mentors)

    final_df = filtered_df[filtered_df["Current Mentor"] == selected_mentor]

    st.write(f"### Students under **{selected_mentor}** ({len(final_df)} students)")
    st.dataframe(final_df, use_container_width=True)

    # prepare single mentor workbook and download
    try:
        mentor_wb = df_to_formatted_workbook(original_wb, sheet_name, final_df, id_col_name="ADEK Applicant ID")
        excel_buffer = BytesIO()
        mentor_wb.save(excel_buffer)
        st.download_button(
            label=f"📥 Download Formatted Excel for {selected_mentor}",
            data=excel_buffer.getvalue(),
            file_name=f"{selected_mentor}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"Error while preparing formatted file: {e}")

    st.write("---")
    st.write("### 📦 Download ZIP (All Mentors)")

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zf:
        for mentor in mentors:
            m_df = filtered_df[filtered_df["Current Mentor"] == mentor]
            try:
                wb = df_to_formatted_workbook(original_wb, sheet_name, m_df, id_col_name="ADEK Applicant ID")
                mb = BytesIO()
                wb.save(mb)
                zf.writestr(f"{mentor} January 2026 Student List.xlsx", mb.getvalue())
            except Exception as e:
                # write a small error text file so you know which mentor failed
                zf.writestr(f"{mentor}_ERROR.txt", f"Failed to build file for {mentor}: {e}")

    st.download_button(
        label="📥 Download ZIP (Formatted Files)",
        data=zip_buffer.getvalue(),
        file_name="Formatted_Mentor_Files.zip",
        mime="application/zip"
    )


